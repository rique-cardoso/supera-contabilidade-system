import os
from datetime import datetime
from django.core.management.base import BaseCommand, CommandError
from django.utils.timezone import make_aware
import openpyxl

# --- AJUSTE OS IMPORTS CONFORME SEU PROJETO ---
from core.models import Usuario
from clientes.models import Cliente, Empresa, EnderecoEmpresa
from processos.models import Processo, FaseProcesso, Vistoria, Taxa
from notificacoes.models import Notificacao


class Command(BaseCommand):
    help = "Lê uma planilha Excel (.xlsx) e importa os dados fictícios para o sistema."

    def add_arguments(self, parser):
        parser.add_argument(
            "caminho_arquivo",
            type=str,
            help="O caminho local para o arquivo .xlsx (ex: dados_teste_sistema.xlsx)",
        )

    def parse_datetime(self, val):
        """Converte strings de data/hora da planilha para objetos datetime cientes do timezone."""
        if not val:
            return None
        if isinstance(val, datetime):
            return make_aware(val)
        try:
            dt = datetime.strptime(str(val).strip(), "%Y-%m-%d %H:%M:%S")
            return make_aware(dt)
        except ValueError:
            try:
                dt = datetime.strptime(str(val).strip(), "%Y-%m-%d")
                return make_aware(dt)
            except ValueError:
                return None

    def parse_date(self, val):
        """Converte o valor para objeto date do Python (usado no data_vencimento)."""
        if not val:
            return None
        if isinstance(val, datetime):
            return val.date()
        try:
            return datetime.strptime(str(val).strip(), "%Y-%m-%d").date()
        except ValueError:
            return None

    def parse_bool(self, val):
        """Converte o valor da célula para booleano de forma segura."""
        if isinstance(val, bool):
            return val
        return str(val).strip().lower() in ["true", "1", "yes", "sim"]

    def handle(self, *args, **options):
        caminho = options["caminho_arquivo"]

        if not os.path.exists(caminho):
            raise CommandError(f"Arquivo não encontrado no caminho: {caminho}")

        self.stdout.write(
            self.style.WARNING(f"Carregando planilha: {caminho}...")
        )
        wb = openpyxl.load_workbook(caminho, data_only=True)

        # Mapeamento para guardar IDs antigos da planilha e IDs novos gerados no banco,
        # prevenindo quebras caso os IDs de autoincremento do banco comecem diferentes.
        map_usuarios = {}
        map_clientes = {}
        map_empresas = {}
        map_processos = {}

        # ----------------------------------------------------
        # 1. USUÁRIOS
        # ----------------------------------------------------
        if "Usuarios" in wb.sheetnames:
            sheet = wb["Usuarios"]
            self.stdout.write("Importando Usuários...")
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Estrutura: ["id", "email", "username", "role", "is_active", "date_joined"]
                id_planilha, email, username, role, is_active, date_joined = (
                    row
                )

                if not email:
                    continue

                # AbstractUser usa create_user para salvar a senha encriptada adequadamente
                user, criado = Usuario.objects.get_or_create(
                    email=email,
                    defaults={
                        "username": username or email.split("@")[0],
                        "role": role or "padrao",
                        "is_active": (
                            self.parse_bool(is_active)
                            if is_active is not None
                            else True
                        ),
                    },
                )
                if criado:
                    user.set_password("teste123")  # Senha padrão de teste
                    if date_joined:
                        user.date_joined = self.parse_datetime(date_joined)
                    user.save()

                map_usuarios[id_planilha] = user
            self.stdout.write(
                self.style.SUCCESS(
                    f"✓ Usuários processados (Total: {len(map_usuarios)})."
                )
            )

        # ----------------------------------------------------
        # 2. CLIENTES
        # ----------------------------------------------------
        if "Clientes" in wb.sheetnames:
            sheet = wb["Clientes"]
            self.stdout.write("Importando Clientes...")
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Estrutura: ["id", "nome_responsavel", "telefone", "email", "cpf", "data_criacao"]
                id_planilha, nome, telefone, email, cpf, data_criacao = row

                if not cpf:
                    continue

                cliente, criado = Cliente.objects.get_or_create(
                    cpf=cpf,
                    defaults={
                        "nome_responsavel": nome,
                        "telefone": telefone,
                        "email": email,
                    },
                )
                map_clientes[id_planilha] = cliente
            self.stdout.write(
                self.style.SUCCESS(
                    f"✓ Clientes processados (Total: {len(map_clientes)})."
                )
            )

        # ----------------------------------------------------
        # 3. EMPRESAS
        # ----------------------------------------------------
        if "Empresas" in wb.sheetnames:
            sheet = wb["Empresas"]
            self.stdout.write("Importando Empresas...")
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Estrutura: ["id", "cliente_id", "nome_empresa", "cnpj", "cnae", "data_criacao"]
                id_planilha, cliente_id, nome, cnpj, cnae, data_criacao = row

                if not cnpj or cliente_id not in map_clientes:
                    continue

                empresa, criado = Empresa.objects.get_or_create(
                    cnpj=cnpj,
                    defaults={
                        "cliente": map_clientes[cliente_id],
                        "nome_empresa": nome,
                        "cnae": str(cnae),
                    },
                )
                map_empresas[id_planilha] = empresa
            self.stdout.write(
                self.style.SUCCESS(
                    f"✓ Empresas processadas (Total: {len(map_empresas)})."
                )
            )

        # ----------------------------------------------------
        # 4. ENDEREÇOS
        # ----------------------------------------------------
        if "Enderecos_Empresas" in wb.sheetnames:
            sheet = wb["Enderecos_Empresas"]
            self.stdout.write("Importando Endereços...")
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Estrutura: ["id", "empresa_id", "logradouro", "numero", "complemento", "bairro", "cidade", "estado", "cep"]
                (
                    _,
                    empresa_id,
                    logradouro,
                    numero,
                    complemento,
                    bairro,
                    cidade,
                    estado,
                    cep,
                ) = row

                if empresa_id not in map_empresas:
                    continue

                EnderecoEmpresa.objects.get_or_create(
                    empresa=map_empresas[empresa_id],
                    defaults={
                        "logradouro": logradouro,
                        "numero": str(numero),
                        "complemento": complemento,
                        "bairro": bairro,
                        "cidade": cidade,
                        "estado": estado,
                        "cep": str(cep),
                    },
                )
            self.stdout.write(self.style.SUCCESS("✓ Endereços processados."))

        # ----------------------------------------------------
        # 5. PROCESSOS
        # ----------------------------------------------------
        if "Processos" in wb.sheetnames:
            sheet = wb["Processos"]
            self.stdout.write("Importando Processos...")
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Estrutura: ["id", "protocolo", "nome", "descricao", "empresa_id", "orgao", "categoria", "status", "data_criacao", "data_vencimento", "licenciamento_ambiental", "criado_por_id"]
                (
                    id_planilha,
                    protocolo,
                    nome,
                    descricao,
                    empresa_id,
                    orgao,
                    categoria,
                    status,
                    data_criacao,
                    data_vencimento,
                    lic_amb,
                    criado_por_id,
                ) = row

                if not protocolo or empresa_id not in map_empresas:
                    continue

                criado_por_user = map_usuarios.get(criado_por_id, None)

                processo, criado = Processo.objects.get_or_create(
                    protocolo=protocolo,
                    defaults={
                        "nome": nome,
                        "descricao": descricao,
                        "empresa": map_empresas[empresa_id],
                        "orgao": orgao,
                        "categoria": categoria or "FUNCIONAMENTO",
                        "status": status or "ATIVO",
                        "data_vencimento": self.parse_date(data_vencimento),
                        "licenciamento_ambiental": self.parse_bool(lic_amb),
                        "criado_por": criado_por_user,
                    },
                )

                # Vincula os responsáveis padrão do processo (M2M)
                if criado and criado_por_user:
                    processo.responsaveis.add(criado_por_user)

                map_processos[id_planilha] = processo
            self.stdout.write(
                self.style.SUCCESS(
                    f"✓ Processos mapeados (Total: {len(map_processos)})."
                )
            )

        # ----------------------------------------------------
        # 6. FASES DOS PROCESSOS
        # ----------------------------------------------------
        if "Fases_Processos" in wb.sheetnames:
            sheet = wb["Fases_Processos"]
            self.stdout.write("Importando Fases/Checklists...")
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Estrutura: ["id", "processo_id", "nome", "is_geral", "ordem", "is_concluido", "data_conclusao", "data_criacao"]
                (
                    _,
                    processo_id,
                    nome,
                    is_geral,
                    ordem,
                    is_concluido,
                    data_conclusao,
                    _,
                ) = row

                if processo_id not in map_processos:
                    continue

                # Nota: Sua model possui unique_together = [['processo', 'nome']]
                FaseProcesso.objects.get_or_create(
                    processo=map_processos[processo_id],
                    nome=nome,
                    defaults={
                        "is_geral": self.parse_bool(is_geral),
                        "ordem": int(ordem or 0),
                        "is_concluido": self.parse_bool(is_concluido),
                        "data_conclusao": self.parse_datetime(data_conclusao),
                    },
                )
            self.stdout.write(self.style.SUCCESS("✓ Fases processadas."))

        # ----------------------------------------------------
        # 7. VISTORIAS
        # ----------------------------------------------------
        if "Vistorias" in wb.sheetnames:
            sheet = wb["Vistorias"]
            self.stdout.write("Importando Vistorias...")
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Estrutura: ["id", "processo_id", "data_hora", "local", "status", "observacoes", "data_criacao"]
                _, processo_id, data_hora, local, status, observacoes, _ = row

                if processo_id not in map_processos:
                    continue

                Vistoria.objects.create(
                    processo=map_processos[processo_id],
                    data_hora=self.parse_datetime(data_hora),
                    local=local,
                    status=status or "AGENDADA",
                    observacoes=observacoes,
                )
            self.stdout.write(self.style.SUCCESS("✓ Vistorias processadas."))

        # ----------------------------------------------------
        # 8. TAXAS
        # ----------------------------------------------------
        if "Taxas" in wb.sheetnames:
            sheet = wb["Taxas"]
            self.stdout.write("Importando Taxas...")
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Estrutura: ["id", "processo_id", "nome", "valor", "is_paga", "is_isento", "data_criacao", "data_pagamento"]
                (
                    _,
                    processo_id,
                    nome,
                    valor,
                    is_paga,
                    is_isento,
                    _,
                    data_pagamento,
                ) = row

                if processo_id not in map_processos:
                    continue

                Taxa.objects.create(
                    processo=map_processos[processo_id],
                    nome=nome,
                    valor=valor,
                    is_paga=self.parse_bool(is_paga),
                    is_isento=self.parse_bool(is_isento),
                    data_pagamento=self.parse_datetime(data_pagamento),
                )
            self.stdout.write(self.style.SUCCESS("✓ Taxas processadas."))

        # ----------------------------------------------------
        # 9. NOTIFICAÇÕES
        # ----------------------------------------------------
        if "Notificacoes" in wb.sheetnames:
            sheet = wb["Notificacoes"]
            self.stdout.write("Importando Notificações...")
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Estrutura: ["id", "processo_id", "titulo", "mensagem", "categoria", "data_geracao", "is_enviada_email", "is_lida"]
                _, processo_id, titulo, mensagem, categoria, _, is_email, is_lida = (
                    row
                )

                if processo_id not in map_processos:
                    continue

                notificacao = Notificacao.objects.create(
                    processo=map_processos[processo_id],
                    titulo=titulo,
                    mensagem=mensagem,
                    categoria=categoria,
                    is_enviada_email=self.parse_bool(is_email),
                    is_lida=self.parse_bool(is_lida),
                )

                # RF24: Vincula os destinatários baseados nos responsáveis do processo
                proc_obj = map_processos[processo_id]
                notificacao.usuarios_destinatarios.set(
                    proc_obj.responsaveis.all()
                )

            self.stdout.write(
                self.style.SUCCESS("✓ Notificações processadas.")
            )

        self.stdout.write(
            self.style.SUCCESS(
                "🎉 Carga de dados fictícios concluída com sucesso!"
            )
        )